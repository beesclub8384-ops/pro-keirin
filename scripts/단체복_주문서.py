from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from collections import defaultdict

wb = Workbook()

members = [
    (1,  "유주현", "청평",  "L"),
    (2,  "이주영", "동광주", "L"),
    (3,  "김웅겸", "김포",  "M"),
    (4,  "김지호", "김포",  "M"),
    (5,  "문신준서","김포", "M"),
    (6,  "이승원", "동서울","L"),
    (7,  "한동현", "동서울","M"),
    (8,  "김태형", "동서울","L"),
    (9,  "김도현", "동서울","M"),
    (10, "강석호", "동서울","M"),
    (11, "김홍일", "개인",  "M"),
    (12, "강민성", "개인",  "L"),
    (13, "이홍주", "개인",  "XL"),
    (14, "최순영", "개인",  "L"),
    (15, "신주헌", "수성",  "S"),
    (16, "김용진", "수성",  "M"),
    (17, "김현경", "충청권","XXL"),
    (18, "송승현", "충청권","XL"),
    (19, "신광호", "충청권","L"),
    (20, "유연종", "충청권","XL"),
    (21, "황인혁", "충청권","XL"),
    (22, "이인우", "-",    "S"),
    (23, "김영규", "-",    "M"),
    (24, "최민호", "-",    "M"),
    (25, "김범수", "-",    "M"),
    (26, "김관희", "-",    "M"),
    (27, "김환윤", "-",    "M"),
    (28, "박성현", "-",    "M"),
    (29, "방극산", "-",    "M"),
    (30, "조주현", "-",    "M"),
    (31, "민선기", "-",    "M"),
    (32, "김민배", "-",    "L"),
    (33, "박준성", "-",    "L"),
    (34, "배석현", "-",    "L"),
    (35, "김영수", "-",    "L"),
    (36, "정환",   "-",    "XL"),
    (37, "정태양", "-",    "XL"),
    (38, "김명섭", "-",    "XL"),
    (39, "이재옥", "-",    "XL"),
    (40, "황준하", "-",    "XL"),
]

def thin_border():
    s = Side(style='thin')
    return Border(left=s, right=s, top=s, bottom=s)

HEADER_FILL = PatternFill("solid", start_color="1F4E79")
HEADER_FONT = Font(name="맑은 고딕", bold=True, color="FFFFFF", size=11)
BODY_FONT   = Font(name="맑은 고딕", size=10)
CENTER      = Alignment(horizontal="center", vertical="center")
LEFT        = Alignment(horizontal="left",   vertical="center")

SIZE_COLORS = {
    "S":   "D9EAD3",
    "M":   "C9DAF8",
    "L":   "FFF2CC",
    "XL":  "FCE5CD",
    "XXL": "F4CCCC",
}

# ── Sheet 1: 전체 명단 ──
ws1 = wb.active
ws1.title = "전체 명단"

ws1.merge_cells("A1:E1")
ws1["A1"].value     = "단체복 주문서 – 전체 명단"
ws1["A1"].font      = Font(name="맑은 고딕", bold=True, size=14, color="1F4E79")
ws1["A1"].alignment = CENTER
ws1.row_dimensions[1].height = 30

for col, (h, w) in enumerate(zip(["번호","이름","소속","사이즈","비고"], [8,14,12,10,18]), 1):
    cell = ws1.cell(row=2, column=col, value=h)
    cell.font = HEADER_FONT; cell.fill = HEADER_FILL
    cell.alignment = CENTER; cell.border = thin_border()
    ws1.column_dimensions[get_column_letter(col)].width = w

for i, (no, name, group, size) in enumerate(members, 3):
    for col, val in enumerate([no, name, group, size, ""], 1):
        cell = ws1.cell(row=i, column=col, value=val)
        cell.font = BODY_FONT; cell.border = thin_border()
        cell.alignment = LEFT if col == 2 else CENTER
        cell.fill = PatternFill("solid", start_color=SIZE_COLORS.get(size, "FFFFFF"))

tr = len(members) + 3
ws1.merge_cells(f"A{tr}:C{tr}")
for col in range(1, 6):
    cell = ws1.cell(row=tr, column=col)
    cell.fill = PatternFill("solid", start_color="D9D9D9")
    cell.border = thin_border()
    cell.font = Font(name="맑은 고딕", bold=True, size=11)
    cell.alignment = CENTER
ws1.cell(row=tr, column=1).value = "합  계"
ws1.cell(row=tr, column=4).value = f'=COUNTA(B3:B{tr-1})&"명"'

# ── Sheet 2: 사이즈별 집계 ──
ws2 = wb.create_sheet("사이즈별 집계")
ws2.merge_cells("A1:C1")
ws2["A1"].value     = "단체복 주문서 – 사이즈별 집계"
ws2["A1"].font      = Font(name="맑은 고딕", bold=True, size=14, color="1F4E79")
ws2["A1"].alignment = CENTER
ws2.row_dimensions[1].height = 30

for col, (h, w) in enumerate(zip(["사이즈","수량","명단"], [12,10,50]), 1):
    cell = ws2.cell(row=2, column=col, value=h)
    cell.font = HEADER_FONT; cell.fill = HEADER_FILL
    cell.alignment = CENTER; cell.border = thin_border()
    ws2.column_dimensions[get_column_letter(col)].width = w

size_map = defaultdict(list)
for _, name, _, size in members:
    size_map[size].append(name)

for i, size in enumerate(["S","M","L","XL","XXL"], 3):
    fill = PatternFill("solid", start_color=SIZE_COLORS.get(size, "FFFFFF"))
    ws2.cell(row=i, column=1, value=size).font      = Font(name="맑은 고딕", bold=True, size=11)
    ws2.cell(row=i, column=1).alignment = CENTER
    ws2.cell(row=i, column=1).border    = thin_border()
    ws2.cell(row=i, column=1).fill      = fill
    ws2.cell(row=i, column=2, value=len(size_map[size])).alignment = CENTER
    ws2.cell(row=i, column=2).border = thin_border()
    ws2.cell(row=i, column=2).fill   = fill
    ws2.cell(row=i, column=2).font   = BODY_FONT
    ws2.cell(row=i, column=3, value=", ".join(size_map[size])).alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    ws2.cell(row=i, column=3).border = thin_border()
    ws2.cell(row=i, column=3).fill   = fill
    ws2.cell(row=i, column=3).font   = Font(name="맑은 고딕", size=9)

tr2 = 8
for col in range(1, 4):
    cell = ws2.cell(row=tr2, column=col)
    cell.fill = PatternFill("solid", start_color="D9D9D9")
    cell.border = thin_border()
    cell.font = Font(name="맑은 고딕", bold=True, size=11)
    cell.alignment = CENTER
ws2.cell(row=tr2, column=1).value = "합  계"
ws2.cell(row=tr2, column=2).value = "=SUM(B3:B7)"

wb.save("단체복_주문서.xlsx")
print("완료!")
