# -*- coding: utf-8 -*-
# 기존 엑셀에 "2024년 전체 출주표" 시트 추가
import json
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side

EXCEL_PATH = Path("C:/Users/win10/Downloads") / "2024년_경주편성_통계분석.xlsx"
JSON_PATH = Path("C:/Users/win10/pro-keirin/scripts") / "2024-entries-export.json"

BIPAUP = set(
    "강민성 강병석 강진원 고재준 고종인 공태민 공태욱 곽현명 구동훈 구본광 "
    "권우주 권혁진 김관희 김광근 김근영 김기동 김다빈 김동관 김동훈 김두용 "
    "김로운 김명래 김명섭 김명중 김민균 김민배 김민수 김민욱 김민준 김민호 "
    "김배영 김범수 김범준 김범중 김시후 김영규 김영석 김영섭 김영수 김옥철 "
    "김용규 김용남 김우겸 김우영 김원진 김원호 김이남 김제영 김종성 김종현 "
    "김주한 김주호 김준빈 김준철 김철민 김태범 김한울 김현 김현경 김형모 "
    "김형완 김홍기 김홍일 김환윤 김희준 남용찬 노태경 노형균 류근철 류재민 "
    "류재열 명경민 문인재 문희덕 민상호 민선기 박경호 박동수 박민철 박석기 "
    "박성순 박성현 박승민 박용범 박종현 박종태 박준성 박진철 박철성 방극산 "
    "배민구 배석현 배준호 석혜윤 성용환 손경수 손성진 손재우 손제용 송경방 "
    "송대호 송승현 송정욱 송종훈 신동현 신은섭 안성민 안창진 양승원 양진우 "
    "엄재천 엄정일 엄희태 여민호 오기호 왕지현 우성식 원신재 원준오 유연종 "
    "유주현 유태복 윤진규 윤현구 윤현준 이규봉 이근우 이기주 이기한 이기호 "
    "이록희 이상현 이서혁 이성록 이성민 이수원 이용희 이우정 이유진 이인우 "
    "이일수 이재봉 이재옥 이정민 이정석 이정운 이지훈 이진웅 이진원 이차현 "
    "이찬우 이태운 이홍주 인치환 임경수 임대성 임유섭 임재연 임채빈 장인석 "
    "전경호 전영규 전원규 정동호 정상민 정윤재 정재원 정정교 정종진 정지민 "
    "정태양 정하늘 정하전 정해권 정해민 조성윤 조영소 조영환 조재호 조주현 "
    "조창인 주성민 주효진 최근영 최대용 최동현 최민호 최병길 최순영 최정환 "
    "한탁희 함동주 함명주 허동혁 현지운 황승호 황인혁 황준하".split()
)

# 스타일
FONT_DEFAULT = Font(name="맑은 고딕", size=10)
FONT_HEADER = Font(name="맑은 고딕", size=10, bold=True, color="FFFFFF")
FONT_BLUE = Font(name="맑은 고딕", size=10, color="1565C0")
FONT_RED = Font(name="맑은 고딕", size=10, color="C62828")
FONT_SOURCE = Font(name="맑은 고딕", size=9, italic=True, color="555555")

FILL_HEADER = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
FILL_SEONBAL = PatternFill(start_color="FFF9C4", end_color="FFF9C4", fill_type="solid")  # 연한 노란
FILL_WOOSOO = PatternFill(start_color="E8F5E9", end_color="E8F5E9", fill_type="solid")   # 연한 초록
FILL_TEUKSEON = PatternFill(start_color="EDE7F6", end_color="EDE7F6", fill_type="solid")  # 연한 보라
FILL_SOURCE = PatternFill(start_color="F0F0F0", end_color="F0F0F0", fill_type="solid")

ALIGN_CENTER = Alignment(horizontal="center", vertical="center")
THIN_BORDER = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"),
)

SOURCE_TEXT = "출처: kcycle.or.kr 출주표 (2024년 광명 경기장 전체 / 분석일: 2026년 3월)"


def get_row_fill(race_type: str):
    if "특선" in race_type:
        return FILL_TEUKSEON
    if "우수" in race_type:
        return FILL_WOOSOO
    if "선발" in race_type:
        return FILL_SEONBAL
    return None


def main():
    with open(str(JSON_PATH), encoding="utf-8") as f:
        data = json.load(f)

    wb = load_workbook(str(EXCEL_PATH))

    # 기존 시트가 있으면 삭제
    if "2024년 전체 출주표" in wb.sheetnames:
        del wb["2024년 전체 출주표"]

    ws = wb.create_sheet(title="2024년 전체 출주표")

    # 열 너비
    col_widths = [8, 8, 8, 14, 10, 10, 10, 10, 10, 10, 10]
    for i, w in enumerate(col_widths, 1):
        from openpyxl.utils import get_column_letter
        ws.column_dimensions[get_column_letter(i)].width = w

    # 출처
    ws.merge_cells("A1:K1")
    c = ws.cell(row=1, column=1, value=SOURCE_TEXT)
    c.font = FONT_SOURCE
    c.fill = FILL_SOURCE
    c.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 20

    # 헤더
    headers = ["회차", "일차", "경주", "경주구분", "1번", "2번", "3번", "4번", "5번", "6번", "7번"]
    for ci, h in enumerate(headers, 1):
        c = ws.cell(row=2, column=ci, value=h)
        c.font = FONT_HEADER
        c.fill = FILL_HEADER
        c.alignment = ALIGN_CENTER
        c.border = THIN_BORDER

    # 데이터
    for ri, race in enumerate(data):
        row = 3 + ri
        row_fill = get_row_fill(race["race_type"])

        vals = [race["round"], race["day"], race["race_no"], race["race_type"]]
        for ci, val in enumerate(vals, 1):
            c = ws.cell(row=row, column=ci, value=val)
            c.font = FONT_DEFAULT
            c.alignment = ALIGN_CENTER
            c.border = THIN_BORDER
            if row_fill:
                c.fill = row_fill

        for bi, name in enumerate(race["names"], 5):
            c = ws.cell(row=row, column=bi, value=name)
            if name in BIPAUP:
                c.font = FONT_BLUE
            elif name:
                c.font = FONT_RED
            else:
                c.font = FONT_DEFAULT
            c.alignment = ALIGN_CENTER
            c.border = THIN_BORDER
            if row_fill:
                c.fill = row_fill

    wb.save(str(EXCEL_PATH))
    print(f"저장 완료: {EXCEL_PATH}")
    print(f"총 {len(data)}경주 추가됨")


if __name__ == "__main__":
    main()
