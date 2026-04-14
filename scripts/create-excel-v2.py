# -*- coding: utf-8 -*-
# 2024년 경주편성 통계분석 엑셀 v2 (결승 제외)
import json
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

JSON_PATH = Path(__file__).resolve().parent / "2024-v2-export.json"
OUT_PATH = Path("C:/Users/win10/Downloads") / "2024년_경주편성_통계분석.xlsx"

BIPAUP = set(
    "강민성 강병석 강진원 고재준 고종인 공태민 공태욱 곽현명 구동훈 구본광 "
    "권우주 권혁진 김관희 김광근 김근영 김기동 김다빈 김동관 김동훈 김두용 "
    "김로운 김명래 김명섭 김명중 김민균 김민배 김민수 김민욱 김민준 김민호 "
    "김배영 김범수 김범준 김범중 김시후 김영규 김영석 김영섭 김영수 김옥철 "
    "김용규 김용남 김우겸 김우영 김원진 김원호 김이남 김제영 김종성 김종현 "
    "김주한 김주호 김준빈 김준철 김철민 김태범 김한울 김현 김현경 김형모 "
    "김형완 김홍기 김홍일 김환윤 김희준 남용찬 노태경 노형균 류근철 류재민 "
    "류재열 명경민 문인재 문희덕 민상호 민선기 박경호 박동수 박민철 박석기 "
    "박성순 박성현 박승민 박용범 박종현 박종태 박준성 박진철 박철성 방극산 "
    "배민구 배석현 배준호 석혜윤 성용환 손경수 손성진 손재우 손제용 송경방 "
    "송대호 송승현 송정욱 송종훈 신동현 신은섭 안성민 안창진 양승원 양진우 "
    "엄재천 엄정일 엄희태 여민호 오기호 왕지현 우성식 원신재 원준오 유연종 "
    "유주현 유태복 윤진규 윤현구 윤현준 이규봉 이근우 이기주 이기한 이기호 "
    "이록희 이상현 이서혁 이성록 이성민 이수원 이용희 이우정 이유진 이인우 "
    "이일수 이재봉 이재옥 이정민 이정석 이정운 이지훈 이진웅 이진원 이차현 "
    "이찬우 이태운 이홍주 인치환 임경수 임대성 임유섭 임재연 임채빈 장인석 "
    "전경호 전영규 전원규 정동호 정상민 정윤재 정재원 정정교 정종진 정지민 "
    "정태양 정하늘 정하전 정해권 정해민 조성윤 조영소 조영환 조재호 조주현 "
    "조창인 주성민 주효진 최근영 최대용 최동현 최민호 최병길 최순영 최정환 "
    "한탁희 함동주 함명주 허동혁 현지운 황승호 황인혁 황준하".split()
)

# ── 스타일 ──
F_DEFAULT = Font(name="맑은 고딕", size=11)
F_TITLE = Font(name="맑은 고딕", size=16, bold=True)
F_SUBTITLE = Font(name="맑은 고딕", size=13, bold=True)
F_HEADER = Font(name="맑은 고딕", size=11, bold=True, color="FFFFFF")
F_BOLD = Font(name="맑은 고딕", size=11, bold=True)
F_SOURCE = Font(name="맑은 고딕", size=10, italic=True, color="555555")
F_NOTE = Font(name="맑은 고딕", size=10, color="C00000")
F_BLUE = Font(name="맑은 고딕", size=10, color="1565C0")
F_RED = Font(name="맑은 고딕", size=10, color="C62828")
F_SHEET = Font(name="맑은 고딕", size=10)

BG_HEADER = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
BG_RED = PatternFill(start_color="FCE4EC", end_color="FCE4EC", fill_type="solid")
BG_BLUE = PatternFill(start_color="E3F2FD", end_color="E3F2FD", fill_type="solid")
BG_YELLOW = PatternFill(start_color="FFF9C4", end_color="FFF9C4", fill_type="solid")
BG_GREEN = PatternFill(start_color="E8F5E9", end_color="E8F5E9", fill_type="solid")
BG_PURPLE = PatternFill(start_color="EDE7F6", end_color="EDE7F6", fill_type="solid")
BG_SOURCE = PatternFill(start_color="F0F0F0", end_color="F0F0F0", fill_type="solid")
BG_TITLE = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
BG_NOTE = PatternFill(start_color="FFF3E0", end_color="FFF3E0", fill_type="solid")

AC = Alignment(horizontal="center", vertical="center")
AL = Alignment(horizontal="left", vertical="center", wrap_text=True)
TB = Border(left=Side("thin"), right=Side("thin"), top=Side("thin"), bottom=Side("thin"))

SOURCE_TEXT = "출처: kcycle.or.kr 출주표 (2024년 광명 경기장 전체 / 분석일: 2026년 3월)"
EXCLUSION_NOTE = (
    "※ 결승 경주(선발결승/우수결승/특선결승/그랑프리결승)는 "
    "편성자 1인의 재량이 아닌 별도 규정에 의해 편성되므로 본 분석에서 제외함. "
    "준결승은 편성자 재량에 해당하므로 포함."
)

DESCS = {
    "선발": {
        0: "파업 선수끼리만 편성된 경주가 비정상적으로 많음",
        1: "비파업 1명 배치가 극단적으로 회피됨",
        2: "기대보다 적음",
        3: "비파업 3명+파업 4명 조합에 비정상적으로 집중",
        4: "기대와 유사",
        5: "기대와 유사",
        6: "기대값 0.4건 대비 10배 (확률적으로 매우 이례적)",
        7: "기대값 0.0건 → 선발급에서 비파업 7명 경주는 확률적으로 불가능한 편성",
    },
    "우수": {
        0: "파업 선수끼리만 편성된 경주가 비정상적으로 많음",
        1: "비파업 1명 배치가 극단적으로 회피됨",
        2: "비파업 2명 배치가 극단적으로 회피됨",
        3: "비파업 3명+파업 4명 조합에 극단적으로 집중",
        4: "기대보다 적음",
        5: "기대보다 극단적으로 적음",
        6: "기대보다 적음",
        7: "비파업끼리만 편성된 경주가 비정상적으로 많음 (기대값 대비 13배)",
    },
    "특선": {
        0: "비파업 0명은 확률적으로 거의 불가능 (기대값 0.1건)",
        1: "기대보다 적음",
        2: "기대보다 적음",
        3: "기대보다 많음",
        4: "비파업 4명+파업 3명 조합에 집중",
        5: "비파업 5명 배치가 극단적으로 회피됨",
        6: "비파업 6명 배치가 극단적으로 회피됨",
        7: "비파업끼리만 편성된 경주가 비정상적으로 많음",
    },
}


def widths(ws, wlist):
    for i, w in enumerate(wlist, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def source_row(ws, row, cols):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=cols)
    c = ws.cell(row=row, column=1, value=SOURCE_TEXT)
    c.font, c.fill, c.alignment = F_SOURCE, BG_SOURCE, AL
    ws.row_dimensions[row].height = 20


def note_row(ws, row, cols):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=cols)
    c = ws.cell(row=row, column=1, value=EXCLUSION_NOTE)
    c.font, c.fill = F_NOTE, BG_NOTE
    c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    ws.row_dimensions[row].height = 36


def header_row(ws, row, headers):
    for ci, h in enumerate(headers, 1):
        c = ws.cell(row=row, column=ci, value=h)
        c.font, c.fill, c.alignment, c.border = F_HEADER, BG_HEADER, AC, TB


def data_cell(ws, row, col, val, font=None, fill=None, align=None):
    c = ws.cell(row=row, column=col, value=val)
    c.font = font or F_DEFAULT
    c.alignment = align or AC
    c.border = TB
    if fill:
        c.fill = fill
    return c


# ── 시트 1: 분석 개요 ──
def build_overview(wb, stats):
    ws = wb.create_sheet("분석 개요")
    widths(ws, [18, 18, 18, 18, 20, 24])

    source_row(ws, 1, 6)
    note_row(ws, 2, 6)

    ws.merge_cells("A4:F4")
    c = ws.cell(row=4, column=1, value="2024년 광명 경기장 경주 편성 통계 분석")
    c.font, c.alignment, c.fill = F_TITLE, AC, BG_TITLE
    ws.row_dimensions[4].height = 36

    ws.merge_cells("A6:F6")
    ws.cell(row=6, column=1, value="분석 목적: 경주 편성이 무작위가 아닌 의도적으로 이루어졌음을 수학적으로 증명").font = F_BOLD

    ws.merge_cells("A7:F7")
    ws.cell(row=7, column=1, value="데이터 출처: kcycle.or.kr 출주표 (2024년 광명 경기장 전체 / 분석일: 2026년 3월)").font = F_DEFAULT

    ws.merge_cells("A9:F9")
    c = ws.cell(row=9, column=1, value="핵심 결론 요약")
    c.font = F_SUBTITLE
    ws.row_dimensions[9].height = 28

    header_row(ws, 11, ["급별", "분석 경주 수", "카이제곱 값", "임계값 (p=0.05)", "임계값 대비 배수", "판정"])

    for ri, grade in enumerate(["선발", "우수", "특선"]):
        s = stats[grade]
        row = 12 + ri
        data_cell(ws, row, 1, grade, F_BOLD)
        data_cell(ws, row, 2, s["raceCount"])
        data_cell(ws, row, 3, f'{s["chi2Total"]:,.2f}', F_BOLD)
        data_cell(ws, row, 4, 14.07)
        data_cell(ws, row, 5, f'{s["chi2Ratio"]}배', F_BOLD)
        data_cell(ws, row, 6, "무작위 편성 불가능", F_BOLD, BG_YELLOW)

    ws.merge_cells("A16:F16")
    ws.cell(row=16, column=1, value="카이제곱 검정: 자유도 7, 유의수준 0.05 기준 임계값 14.07. 이 값을 초과하면 무작위 가설을 기각합니다.").font = F_DEFAULT
    ws.merge_cells("A17:F17")
    ws.cell(row=17, column=1, value="세 급별 모두 임계값의 91~654배에 달하는 카이제곱 값이 나왔으며, 이 결과가 우연히 발생할 확률은 사실상 0%입니다.").font = F_BOLD


# ── 시트 2~4: 급별 상세 ──
def build_detail(wb, grade, stats):
    s = stats[grade]
    title = f"{grade}급 상세"
    ws = wb.create_sheet(title)
    widths(ws, [28, 20, 20, 18, 55])

    source_row(ws, 1, 5)
    note_row(ws, 2, 5)

    ws.merge_cells("A4:E4")
    c = ws.cell(row=4, column=1, value=f'{grade}급 경주 편성 분석 ({s["raceCount"]}경주)')
    c.font, c.alignment, c.fill = F_TITLE, AC, BG_TITLE
    ws.row_dimensions[4].height = 36

    ws.merge_cells("A6:E6")
    desc = (
        f'전체 슬롯: {s["totalSlots"]:,}건 / 비파업 슬롯: {s["bpSlots"]:,}건 / '
        f'비파업 비율(p): {s["pPct"]}%'
    )
    ws.cell(row=6, column=1, value=desc).font = F_BOLD

    ws.merge_cells("A7:E7")
    ws.cell(row=7, column=1, value=(
        "만약 편성자가 비파업/파업 여부를 전혀 고려하지 않았다면, "
        f'경주당 비파업 선수 수는 수학적 확률({s["pPct"]}%)에 따라 분포해야 합니다.'
    )).font = F_DEFAULT

    header_row(ws, 9, ["경주 구성", "무작위일 때 기대 경주 수", "실제 경주 수", "차이 (실제-기대)", "설명"])

    for d in s["dist"]:
        k = d["k"]
        row = 10 + k
        diff = d["diff"]
        diff_str = f'+{diff}' if diff > 0 else str(diff)
        fill = BG_RED if diff > 0 else BG_BLUE if diff < 0 else None

        data_cell(ws, row, 1, f'비파업 {k}명 + 파업 {7-k}명', fill=fill, align=AL)
        data_cell(ws, row, 2, d["expected"], fill=fill)
        data_cell(ws, row, 3, d["actual"], F_BOLD, fill=fill)
        data_cell(ws, row, 4, diff_str, F_BOLD, fill=fill)
        data_cell(ws, row, 5, DESCS[grade].get(k, ""), fill=fill, align=AL)

    cr = 19
    ws.merge_cells(start_row=cr, start_column=1, end_row=cr, end_column=5)
    c = ws.cell(row=cr, column=1, value=(
        f'카이제곱 검정 결과: {s["chi2Total"]:,.2f} '
        f'(임계값 14.07의 {s["chi2Ratio"]}배) → '
        f'이 차이가 우연히 발생할 확률은 사실상 0%입니다'
    ))
    c.font, c.fill, c.alignment, c.border = F_BOLD, BG_YELLOW, AC, TB
    ws.row_dimensions[cr].height = 30


# ── 시트 5: 3개 급 비교 ──
def build_comparison(wb, stats):
    ws = wb.create_sheet("3개 급 비교")
    widths(ws, [14, 14, 14, 14, 14, 14, 14, 14, 14, 14])

    source_row(ws, 1, 10)
    note_row(ws, 2, 10)

    ws.merge_cells("A4:J4")
    c = ws.cell(row=4, column=1, value="선발 / 우수 / 특선 급별 비파업 분포 비교")
    c.font, c.alignment, c.fill = F_TITLE, AC, BG_TITLE
    ws.row_dimensions[4].height = 36

    # 기본 정보 비교
    header_row(ws, 6, ["항목", "선발", "우수", "특선", "", "", "", "", "", ""])
    info_labels = ["분석 경주 수", "전체 슬롯", "비파업 슬롯", "비파업 비율", "카이제곱", "임계값 대비"]
    for ri, label in enumerate(info_labels):
        row = 7 + ri
        data_cell(ws, row, 1, label, F_BOLD)
        for ci, grade in enumerate(["선발", "우수", "특선"], 2):
            s = stats[grade]
            if label == "분석 경주 수": val = f'{s["raceCount"]}경주'
            elif label == "전체 슬롯": val = f'{s["totalSlots"]:,}'
            elif label == "비파업 슬롯": val = f'{s["bpSlots"]:,}'
            elif label == "비파업 비율": val = f'{s["pPct"]}%'
            elif label == "카이제곱": val = f'{s["chi2Total"]:,.2f}'
            else: val = f'{s["chi2Ratio"]}배'
            data_cell(ws, row, ci, val)

    # 분포 비교
    row_start = 15
    ws.merge_cells(start_row=row_start, start_column=1, end_row=row_start, end_column=10)
    ws.cell(row=row_start, column=1, value="비파업 인원별 경주 수 비교 (기대 vs 실제)").font = F_SUBTITLE

    header_row(ws, row_start + 2, [
        "비파업", "선발 기대", "선발 실제", "선발 차이",
        "우수 기대", "우수 실제", "우수 차이",
        "특선 기대", "특선 실제", "특선 차이",
    ])

    for k in range(8):
        row = row_start + 3 + k
        data_cell(ws, row, 1, f"{k}명", F_BOLD)
        for gi, grade in enumerate(["선발", "우수", "특선"]):
            d = stats[grade]["dist"][k]
            diff = d["diff"]
            diff_str = f'+{diff}' if diff > 0 else str(diff)
            fill = BG_RED if diff > 0 else BG_BLUE if diff < 0 else None
            base = 2 + gi * 3
            data_cell(ws, row, base, d["expected"], fill=fill)
            data_cell(ws, row, base + 1, d["actual"], F_BOLD, fill=fill)
            data_cell(ws, row, base + 2, diff_str, F_BOLD, fill=fill)


# ── 시트 6: 전체 출주표 ──
def build_entries(wb, races_data):
    ws = wb.create_sheet("2024년 전체 출주표")
    widths(ws, [8, 8, 8, 14, 10, 10, 10, 10, 10, 10, 10])

    source_row(ws, 1, 11)
    note_row(ws, 2, 11)

    headers = ["회차", "일차", "경주", "경주구분", "1번", "2번", "3번", "4번", "5번", "6번", "7번"]
    header_row(ws, 3, headers)

    def row_fill(rt):
        if "특선" in rt: return BG_PURPLE
        if "우수" in rt: return BG_GREEN
        if "선발" in rt: return BG_YELLOW
        return None

    for ri, race in enumerate(races_data):
        row = 4 + ri
        rf = row_fill(race["race_type"])
        for ci, val in enumerate([race["round"], race["day"], race["race_no"], race["race_type"]], 1):
            data_cell(ws, row, ci, val, F_SHEET, rf)
        for bi, name in enumerate(race["names"], 5):
            if name in BIPAUP:
                font = F_BLUE
            elif name:
                font = F_RED
            else:
                font = F_SHEET
            data_cell(ws, row, bi, name, font, rf)


def main():
    # 기존 파일 삭제
    if OUT_PATH.exists():
        OUT_PATH.unlink()
        print(f"기존 파일 삭제: {OUT_PATH}")

    with open(str(JSON_PATH), encoding="utf-8") as f:
        data = json.load(f)

    stats = data["stats"]
    races = data["races"]

    wb = Workbook()
    wb.remove(wb.active)

    build_overview(wb, stats)
    build_detail(wb, "선발", stats)
    build_detail(wb, "우수", stats)
    build_detail(wb, "특선", stats)
    build_comparison(wb, stats)
    build_entries(wb, races)

    wb.save(str(OUT_PATH))
    print(f"저장 완료: {OUT_PATH}")
    print(f"시트: {wb.sheetnames}")
    for g in ["선발", "우수", "특선"]:
        s = stats[g]
        print(f"  {g}: {s['raceCount']}경주, p={s['pPct']}%, chi2={s['chi2Total']} ({s['chi2Ratio']}배)")


if __name__ == "__main__":
    main()
