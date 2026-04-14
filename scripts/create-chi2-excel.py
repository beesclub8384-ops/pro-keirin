# -*- coding: utf-8 -*-
# 2024년 경주편성 통계분석 엑셀 생성
import json
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

OUT_PATH = Path("C:/Users/win10/Downloads") / "2024년_경주편성_통계분석.xlsx"

# 스타일 정의
FONT_DEFAULT = Font(name="맑은 고딕", size=11)
FONT_TITLE = Font(name="맑은 고딕", size=16, bold=True)
FONT_SUBTITLE = Font(name="맑은 고딕", size=13, bold=True)
FONT_HEADER = Font(name="맑은 고딕", size=11, bold=True, color="FFFFFF")
FONT_BOLD = Font(name="맑은 고딕", size=11, bold=True)
FONT_SOURCE = Font(name="맑은 고딕", size=10, italic=True, color="555555")
FONT_CONCLUSION = Font(name="맑은 고딕", size=11, bold=True)
FONT_DESC = Font(name="맑은 고딕", size=11)

FILL_HEADER = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
FILL_RED_LIGHT = PatternFill(start_color="FCE4EC", end_color="FCE4EC", fill_type="solid")
FILL_BLUE_LIGHT = PatternFill(start_color="E3F2FD", end_color="E3F2FD", fill_type="solid")
FILL_YELLOW = PatternFill(start_color="FFF9C4", end_color="FFF9C4", fill_type="solid")
FILL_SOURCE = PatternFill(start_color="F0F0F0", end_color="F0F0F0", fill_type="solid")
FILL_TITLE_BG = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")

ALIGN_CENTER = Alignment(horizontal="center", vertical="center")
ALIGN_LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)
ALIGN_RIGHT = Alignment(horizontal="right", vertical="center")

THIN_BORDER = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"),
)

SOURCE_TEXT = "출처: kcycle.or.kr 출주표 (2024년 광명 경기장 전체 / 분석일: 2026년 3월)"


def set_col_widths(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def write_source_row(ws, row, col_end):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=col_end)
    c = ws.cell(row=row, column=1, value=SOURCE_TEXT)
    c.font = FONT_SOURCE
    c.fill = FILL_SOURCE
    c.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[row].height = 22


def write_header_row(ws, row, headers):
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=row, column=col, value=h)
        c.font = FONT_HEADER
        c.fill = FILL_HEADER
        c.alignment = ALIGN_CENTER
        c.border = THIN_BORDER


def build_overview(wb):
    ws = wb.create_sheet(title="분석 개요")
    set_col_widths(ws, [18, 18, 18, 18, 20, 24])

    # 출처
    write_source_row(ws, 1, 6)

    # 제목
    ws.merge_cells("A3:F3")
    c = ws.cell(row=3, column=1, value="2024년 광명 경기장 경주 편성 통계 분석")
    c.font = FONT_TITLE
    c.alignment = Alignment(horizontal="center", vertical="center")
    c.fill = FILL_TITLE_BG
    ws.row_dimensions[3].height = 36

    # 분석 목적
    ws.merge_cells("A5:F5")
    c = ws.cell(row=5, column=1, value="분석 목적: 경주 편성이 무작위가 아닌 의도적으로 이루어졌음을 수학적으로 증명")
    c.font = FONT_BOLD
    c.alignment = ALIGN_LEFT

    # 데이터 출처
    ws.merge_cells("A6:F6")
    c = ws.cell(row=6, column=1, value="데이터 출처: kcycle.or.kr 출주표 (2024년 광명 경기장 전체 / 분석일: 2026년 3월)")
    c.font = FONT_DESC
    c.alignment = ALIGN_LEFT

    # 핵심 결론 요약
    ws.merge_cells("A8:F8")
    c = ws.cell(row=8, column=1, value="핵심 결론 요약")
    c.font = FONT_SUBTITLE
    c.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[8].height = 28

    headers = ["급별", "전체 경주 수", "카이제곱 값", "임계값 (p=0.05)", "임계값 대비 배수", "판정"]
    write_header_row(ws, 10, headers)

    data = [
        ["선발", 765, 7341, 14.07, "522배", "무작위 편성 불가능"],
        ["우수", 1071, 1185, 14.07, "84배", "무작위 편성 불가능"],
        ["특선", 652, 3178, 14.07, "226배", "무작위 편성 불가능"],
    ]
    for ri, row_data in enumerate(data):
        row = 11 + ri
        for ci, val in enumerate(row_data):
            c = ws.cell(row=row, column=ci + 1, value=val)
            c.font = FONT_BOLD if ci >= 4 else FONT_DEFAULT
            c.alignment = ALIGN_CENTER
            c.border = THIN_BORDER
            if ci == 5:
                c.fill = FILL_YELLOW
                c.font = FONT_CONCLUSION

    # 설명
    ws.merge_cells("A15:F15")
    c = ws.cell(row=15, column=1, value="카이제곱 검정: 자유도 7, 유의수준 0.05 기준 임계값 14.07. 이 값을 초과하면 무작위 가설을 기각합니다.")
    c.font = FONT_DESC
    c.alignment = ALIGN_LEFT

    ws.merge_cells("A16:F16")
    c = ws.cell(row=16, column=1, value="세 급별 모두 임계값의 84~522배에 달하는 카이제곱 값이 나왔으며, 이 결과가 우연히 발생할 확률은 사실상 0%입니다.")
    c.font = FONT_BOLD
    c.alignment = ALIGN_LEFT


def build_detail_sheet(wb, title, sheet_title, race_count, prob_pct, data_rows, chi2_val, chi2_ratio):
    ws = wb.create_sheet(title=title)
    set_col_widths(ws, [28, 20, 20, 18, 50])

    # 출처
    write_source_row(ws, 1, 5)

    # 제목
    ws.merge_cells("A3:E3")
    c = ws.cell(row=3, column=1, value=sheet_title)
    c.font = FONT_TITLE
    c.alignment = Alignment(horizontal="center", vertical="center")
    c.fill = FILL_TITLE_BG
    ws.row_dimensions[3].height = 36

    # 설명문
    ws.merge_cells("A5:E5")
    desc = (
        f"만약 편성자가 비파업/파업 여부를 전혀 고려하지 않았다면, "
        f"경주당 비파업 선수 수는 수학적 확률({prob_pct}%)에 따라 분포해야 합니다. "
        f"아래 표는 그 기대값과 실제값을 비교합니다."
    )
    c = ws.cell(row=5, column=1, value=desc)
    c.font = FONT_DESC
    c.alignment = ALIGN_LEFT
    ws.row_dimensions[5].height = 40

    # 헤더
    headers = ["경주 구성", "무작위일 때 기대 경주 수", "실제 경주 수", "차이 (실제-기대)", "설명"]
    write_header_row(ws, 7, headers)

    # 데이터
    descriptions = {
        "선발": [
            "기대보다 190건 많음 → 파업 선수끼리만 편성된 경주가 비정상적으로 많음",
            "기대보다 233건 적음 → 비파업 1명+파업 6명 조합이 극단적으로 회피됨",
            "기대보다 56건 적음",
            "기대보다 89건 많음 → 비파업 3명+파업 4명 조합이 비정상적으로 많음",
            "기대와 유사",
            "기대와 유사",
            "기대보다 3.6건 많음 (기대값 0.4건 대비 10배)",
            "기대보다 11건 많음 (기대값 0.0건) → 선발급에서 비파업 7명 경주는 확률적으로 불가능",
        ],
        "우수": [
            "기대보다 55건 많음 → 파업 선수끼리만 편성된 경주가 비정상적으로 많음",
            "기대보다 89건 적음 → 비파업 1명 배치가 극단적으로 회피됨",
            "기대보다 155건 적음 → 비파업 2명 배치가 극단적으로 회피됨",
            "기대보다 350건 많음 → 비파업 3명+파업 4명 조합에 극단적으로 집중",
            "기대보다 91건 적음",
            "기대보다 92건 적음",
            "기대보다 11건 적음",
            "기대보다 34건 많음 (기대값 3.2건 대비 12배) → 비파업끼리만 편성된 경주가 비정상적으로 많음",
        ],
        "특선": [
            "기대보다 16건 많음 (기대값 0.1건) → 특선급에서 비파업 0명은 확률적으로 거의 불가능",
            "기대보다 1.7건 적음",
            "기대보다 3.9건 적음",
            "기대보다 24건 많음",
            "기대보다 84건 많음 → 비파업 4명+파업 3명 조합에 집중",
            "기대보다 170건 적음 → 비파업 5명 배치가 극단적으로 회피됨",
            "기대보다 87건 적음 → 비파업 6명 배치가 극단적으로 회피됨",
            "기대보다 139건 많음 → 비파업끼리만 편성된 경주가 비정상적으로 많음",
        ],
    }

    for ri, (bp_count, expected, actual, diff) in enumerate(data_rows):
        row = 8 + ri
        paup_count = 7 - bp_count
        composition = f"비파업 {bp_count}명 + 파업 {paup_count}명"
        diff_str = f"+{diff}" if diff > 0 else str(diff)

        ws.cell(row=row, column=1, value=composition).font = FONT_DEFAULT
        ws.cell(row=row, column=2, value=expected).font = FONT_DEFAULT
        ws.cell(row=row, column=3, value=actual).font = FONT_BOLD
        ws.cell(row=row, column=4, value=diff_str).font = FONT_BOLD
        ws.cell(row=row, column=5, value=descriptions[title][ri]).font = FONT_DEFAULT

        for ci in range(1, 6):
            c = ws.cell(row=row, column=ci)
            c.border = THIN_BORDER
            if ci <= 4:
                c.alignment = ALIGN_CENTER
            else:
                c.alignment = ALIGN_LEFT

            # 배경색: 차이 양수=빨간, 음수=파란
            if diff > 0:
                c.fill = FILL_RED_LIGHT
            elif diff < 0:
                c.fill = FILL_BLUE_LIGHT

    # 카이제곱 결론
    conclusion_row = 8 + len(data_rows) + 1
    ws.merge_cells(start_row=conclusion_row, start_column=1, end_row=conclusion_row, end_column=5)
    c = ws.cell(
        row=conclusion_row, column=1,
        value=f"카이제곱 검정 결과: {chi2_val:,} (임계값 14.07의 {chi2_ratio}배) → 이 차이가 우연히 발생할 확률은 사실상 0%입니다"
    )
    c.font = FONT_CONCLUSION
    c.fill = FILL_YELLOW
    c.alignment = Alignment(horizontal="center", vertical="center")
    c.border = THIN_BORDER
    ws.row_dimensions[conclusion_row].height = 30


def main():
    wb = Workbook()
    wb.remove(wb.active)

    build_overview(wb)

    # 선발급
    build_detail_sheet(wb, "선발", "선발급 경주 편성 분석 (765경주)", 765, "21.79", [
        (0, 136.9, 327, 190.1),
        (1, 267.0, 34, -233.0),
        (2, 223.2, 167, -56.2),
        (3, 103.7, 193, 89.3),
        (4, 28.9, 26, -2.9),
        (5, 4.8, 3, -1.8),
        (6, 0.4, 4, 3.6),
        (7, 0.0, 11, 11.0),
    ], 7341, 522)

    # 우수급
    build_detail_sheet(wb, "우수", "우수급 경주 편성 분석 (1,071경주)", 1071, "43.59", [
        (0, 19.5, 74, 54.5),
        (1, 105.3, 16, -89.3),
        (2, 244.1, 89, -155.1),
        (3, 314.4, 664, 349.6),
        (4, 242.9, 152, -90.9),
        (5, 112.6, 21, -91.6),
        (6, 29.0, 18, -11.0),
        (7, 3.2, 37, 33.8),
    ], 1185, 84)

    # 특선급
    build_detail_sheet(wb, "특선", "특선급 경주 편성 분석 (652경주)", 652, "71.64", [
        (0, 0.1, 16, 15.9),
        (1, 1.7, 0, -1.7),
        (2, 12.9, 9, -3.9),
        (3, 54.3, 78, 23.7),
        (4, 137.1, 221, 83.9),
        (5, 207.8, 38, -169.8),
        (6, 174.9, 88, -86.9),
        (7, 63.1, 202, 138.9),
    ], 3178, 226)

    wb.save(str(OUT_PATH))
    print(f"저장 완료: {OUT_PATH}")


if __name__ == "__main__":
    main()
