# 연도별 노조별 출전선수 명단 엑셀 생성
# 입력: scripts/union-compare-result.json
# 출력: C:/Users/win10/Downloads/연도별_노조별_출전선수명단.xlsx

import json
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

SCRIPT_DIR = Path(__file__).resolve().parent
JSON_PATH = SCRIPT_DIR / "union-compare-result.json"
OUT_PATH = Path("C:/Users/win10/Downloads/연도별_노조별_출전선수명단.xlsx")

SOURCE_TEXT = (
    "출처: kcycle.or.kr 출주표 (수집일: 2026년 3월 기준) "
    "/ 프로경륜선수노동조합 명단 (2025년 10월 2일 설립 기준)"
)

# 헤더 스타일
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
HEADER_ALIGN = Alignment(horizontal="center")

BLUE_FILL = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
GREEN_FILL = PatternFill(start_color="548235", end_color="548235", fill_type="solid")
RED_FILL = PatternFill(start_color="C00000", end_color="C00000", fill_type="solid")

SOURCE_FILL = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
SOURCE_FONT = Font(italic=True, size=10, color="333333")

TOTAL_FONT = Font(bold=True, size=11)

HEADERS = [
    ("비파업_프로경륜", BLUE_FILL, "pro"),
    ("비파업_비조합원", GREEN_FILL, "non"),
    ("파업_한국경륜", RED_FILL, "han"),
]


def build_sheet(ws, year_data):
    """시트 하나를 구성한다."""

    # ── 1행: 출처 (A1:C1 병합) ──
    ws.merge_cells("A1:C1")
    cell = ws["A1"]
    cell.value = SOURCE_TEXT
    cell.font = SOURCE_FONT
    cell.fill = SOURCE_FILL
    cell.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 22

    # ── 3행: 헤더 ──
    for col_idx, (title, fill, _) in enumerate(HEADERS, start=1):
        cell = ws.cell(row=3, column=col_idx, value=title)
        cell.font = HEADER_FONT
        cell.fill = fill
        cell.alignment = HEADER_ALIGN

    # ── 4행~: 선수 이름 ──
    max_rows = max(len(year_data[key]) for _, _, key in HEADERS)

    for col_idx, (_, _, key) in enumerate(HEADERS, start=1):
        names = year_data[key]
        for row_offset, name in enumerate(names):
            ws.cell(row=4 + row_offset, column=col_idx, value=name)

    # ── 합계행 ──
    total_row = 4 + max_rows + 1  # 한 줄 띄우고
    for col_idx, (_, _, key) in enumerate(HEADERS, start=1):
        col_letter = get_column_letter(col_idx)
        # COUNTA 수식: 데이터 범위 4행 ~ 4+max_rows-1행
        data_end = 4 + max_rows - 1
        cell = ws.cell(
            row=total_row,
            column=col_idx,
            value=f'=COUNTA({col_letter}4:{col_letter}{data_end})',
        )
        cell.font = TOTAL_FONT
        cell.alignment = Alignment(horizontal="center")

    # ── 열 너비 ──
    for col_idx in range(1, 4):
        ws.column_dimensions[get_column_letter(col_idx)].width = 15


def main():
    with open(JSON_PATH, encoding="utf-8") as f:
        data = json.load(f)

    wb = Workbook()
    # 기본 시트 제거
    wb.remove(wb.active)

    for year in ["2022", "2023", "2024", "2025"]:
        ws = wb.create_sheet(title=year)
        build_sheet(ws, data[year])

    wb.save(str(OUT_PATH))
    print(f"저장 완료: {OUT_PATH}")
    for year in ["2022", "2023", "2024", "2025"]:
        d = data[year]
        print(f"  {year}: 프로경륜={len(d['pro'])}명, 비조합원={len(d['non'])}명, 한국경륜={len(d['han'])}명")


if __name__ == "__main__":
    main()
